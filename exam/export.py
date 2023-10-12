from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook

from .filters import DaftarFilter
from .models import Daftar

harini = timezone.now().strftime('%d-%m-%Y')


@login_required
def export_xls(request):
    response = HttpResponse(content_type='application/ms-excel')
    namafile = f"bcs_{harini}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={namafile}'
    print(request.GET)
    param = request.GET.copy()
    parameter = param.pop("page", True) and param.urlencode()  # buang page dari url

    wb = Workbook()
    ws = wb.active
    ws.title = 'BCS'

    row_num = 0

    columns = ['Tarikh', 'No. BCS', 'MRN', 'NRIC', 'Nama', 'Wad', 'Modaliti', 'Exam', 'Merge', 'Catatan', 'JXR','Merged Oleh']

    for col_num in range(0, len(columns)):
        ws.cell(row=1, column=col_num + 1, value=columns[col_num])

    # Sheet body, remaining rows

    daftar = DaftarFilter(
        request.GET,
        queryset=Daftar.objects.all()
        .select_related("bcs__ward", "exam__modaliti")
        .order_by("-bcs__tarikh").values_list('bcs__tarikh', 'nobcs', 'bcs__mrn', 'bcs__nric', 'bcs__nama',
                                              'bcs__ward__wad', 'exam__modaliti__nama', 'exam__exam', 'merged',
                                              'dcatatan', 'jxr__nama','merge_by__nama'),
    )
    rows = list(daftar.qs)
    row_num = 0
    for row in rows:
        row_num += 1
        for col_num in range(0, len(row)):
            ws.cell(row=row_num + 1, column=col_num + 1, value=str(row[col_num]))

    wb.save(response)
    return response
